import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
import os
import tkinter.messagebox as messagebox


def highlight_excel_rows(
    file_to_highlight_path: str,
    file_to_highlight_sheet_name: str,
    data_file_path: str,
    data_sheet_name: str,
    highlight_color: str,
    status_callback,
):
    status_callback("--- Starting Excel Row Highlighting ---")

    comparison_key_cols = ["Name", "CVE", "Host (IP address)", "Port"]

    # normalize hex (support both 'ffff00' and '#ffff00')
    if highlight_color.startswith('#'):
        clean_color = highlight_color[1:]
    else:
        clean_color = highlight_color

    highlight_fill = PatternFill(start_color=clean_color, end_color=clean_color, fill_type="solid")

    try:
        status_callback(f"Loading file to highlight: '{file_to_highlight_path}' (Sheet: '{file_to_highlight_sheet_name}')")
        df_to_highlight = pd.read_excel(file_to_highlight_path, sheet_name=file_to_highlight_sheet_name, dtype=str)
        df_to_highlight.fillna("", inplace=True)

        status_callback(f"Loading data file for comparison: '{data_file_path}' (Sheet: '{data_sheet_name}')")
        df_data = pd.read_excel(data_file_path, sheet_name=data_sheet_name, dtype=str)
        df_data.fillna("", inplace=True)
        status_callback("Excel files loaded successfully for highlighting.")

    except FileNotFoundError as e:
        status_callback(f"Error: One of the files was not found. {e}")
        messagebox.showerror("File Error", f"One of the files was not found: {e}")
        return
    except ValueError as e:
        status_callback(f"Error: Sheet name not found. {e}")
        messagebox.showerror("Sheet Error", f"Sheet name not found: {e}")
        return
    except Exception as e:
        status_callback(f"An unexpected error occurred while loading Excel files for highlighting: {e}")
        messagebox.showerror("Loading Error", f"An unexpected error occurred: {e}")
        return

    # validate columns exist in both dataframes
    for df_name, df, file_path in [("File to Highlight", df_to_highlight, file_to_highlight_path), ("Data File", df_data, data_file_path)]:
        for col in comparison_key_cols:
            if col not in df.columns:
                error_msg = f"Error: The '{df_name}' ({os.path.basename(file_path)}) is missing a crucial comparison column: '{col}'."
                status_callback(error_msg)
                messagebox.showerror("Column Error", error_msg)
                return

    status_callback("Generating keys for comparison...")
    data_keys_set = set(
        tuple((str(row[col]) if pd.notna(row[col]) else '').strip() for col in comparison_key_cols)
        for _, row in df_data.iterrows()
    )

    status_callback(f"Opening '{file_to_highlight_path}' for highlighting...")
    try:
        wb = openpyxl.load_workbook(file_to_highlight_path)
        if file_to_highlight_sheet_name not in wb.sheetnames:
            status_callback(f"Error: Sheet '{file_to_highlight_sheet_name}' not found in '{file_to_highlight_path}'.")
            messagebox.showerror("Sheet Error", f"Sheet '{file_to_highlight_sheet_name}' not found.")
            return
        ws = wb[file_to_highlight_sheet_name]

        highlighted_count = 0
        total_rows = len(df_to_highlight)
        for idx, row in df_to_highlight.iterrows():
            key = tuple((str(row[col]) if pd.notna(row[col]) else '').strip() for col in comparison_key_cols)
            if key in data_keys_set:
                # highlight the row
                for col_idx in range(1, ws.max_column + 1):
                    ws.cell(row=idx + 2, column=col_idx).fill = highlight_fill
                highlighted_count += 1

            percent = int(((idx + 1) / total_rows) * 100) if total_rows > 0 else 100
            status_callback(f"Processing... {percent}% complete ({idx + 1}/{total_rows} rows)")

        wb.save(file_to_highlight_path)
        status_callback(f"✅ Highlighting complete! {highlighted_count} rows highlighted.")
        messagebox.showinfo("Success", f"Matching rows are highlighted in '{os.path.basename(file_to_highlight_path)}'!")

    except Exception as e:
        status_callback(f"An error occurred while highlighting: {e}")
        messagebox.showerror("Highlighting Error", str(e))
