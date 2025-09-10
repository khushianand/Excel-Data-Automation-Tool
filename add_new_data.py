import pandas as pd
import os
import tkinter.messagebox as messagebox
import openpyxl
from openpyxl.styles import PatternFill


def add_new_data_to_tracking(new_data_file_path, new_data_sheet_name, tracking_file_path, tracking_sheet_name, highlight_color_hex, status_callback=None): # MODIFIED
    try:
        if status_callback:
            status_callback("--- Starting 'Add New Data' Process ---")

        df_new_data = pd.read_excel(new_data_file_path, sheet_name=new_data_sheet_name) # MODIFIED
        if df_new_data.empty:
            messagebox.showinfo("Info", "The new data file is empty. Nothing to add.")
            if status_callback:
                status_callback("No rows to append (new data file empty).")
            return

    except FileNotFoundError:
        messagebox.showerror("File Error", f"Could not read new data file: {os.path.basename(new_data_file_path)}")
        if status_callback:
            status_callback("File not found for new data.")
        return
    except Exception as e:
        messagebox.showerror("File Error", f"Could not read new data from '{os.path.basename(new_data_file_path)}'. Error: {e}")
        if status_callback:
            status_callback(f"Error reading new data: {e}")
        return

    # Open tracking workbook and ensure sheet exists
    try:
        wb = openpyxl.load_workbook(tracking_file_path)
        if tracking_sheet_name not in wb.sheetnames:
            messagebox.showerror("Sheet Error", f"Sheet '{tracking_sheet_name}' not found in the tracking file.")
            if status_callback:
                status_callback(f"Sheet '{tracking_sheet_name}' not found in tracking workbook.")
            return
        ws = wb[tracking_sheet_name]

    except FileNotFoundError:
        messagebox.showerror("File Error", f"Tracking file not found: {os.path.basename(tracking_file_path)}")
        if status_callback:
            status_callback("Tracking file not found.")
        return
    except Exception as e:
        messagebox.showerror("File Error", f"Could not open tracking workbook: {e}")
        if status_callback:
            status_callback(f"Error opening tracking workbook: {e}")
        return

    # Prepare highlight fill
    clean_color = highlight_color_hex[1:] if highlight_color_hex.startswith('#') else highlight_color_hex
    fill = PatternFill(start_color=clean_color, end_color=clean_color, fill_type="solid")

    # Determine where to start appending (next free row)
    start_row = ws.max_row + 1

    # Append rows from df_new_data (preserve column order if possible)
    try:
        cols = list(df_new_data.columns)
        for r_idx, row in df_new_data.iterrows():
            for c_idx, col in enumerate(cols, start=1):
                value = row[col]
                ws.cell(row=start_row + r_idx, column=c_idx, value=value)

        # Apply highlighting to appended rows
        for row_num in range(start_row, start_row + len(df_new_data)):
            for cell in ws[row_num]:
                cell.fill = fill

        wb.save(tracking_file_path)

        messagebox.showinfo("Success", f"Successfully added {len(df_new_data)} rows to the '{tracking_sheet_name}' sheet and highlighted them.")
        if status_callback:
            status_callback(f"✅ Appended {len(df_new_data)} rows to '{tracking_sheet_name}' and highlighted them.")

    except Exception as e:
        messagebox.showerror("Error", f"An unexpected error occurred while appending data: {e}")
        if status_callback:
            status_callback(f"Error appending data: {e}")